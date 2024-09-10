import openpyxl as xl
from openpyxl.chart import BarChart, Reference #these are classes

#main point of the whole project is to update the prices
#the price column in the excel file(3rd column) was wrong
#each price needs to be decreased by 10%
#also create a chart of the data

#read the excel file
#db=pd.read_excel('example.xlsx')
def process_workbook(filename):
   #Load the excel spreadsheet and return a work book object
   wb=xl.load_workbook(filename)

   #Returns the first sheet
   sheet=wb['Sheet1'] #case sensitive use big S for Sheet

   #get back the transaction_id cell [Row A, column 1]
   cell=sheet.cell(1,1) 

   #cell=sheet['a1'] is equivalent to above


   for row in range(2, sheet.max_row+1): #start from two to ignore the names of the first row 
      cell=sheet.cell(row,3) #will go through the whole 3rd column
      corrected_price=cell.value * 0.9
      corrected_price_cell=sheet.cell(row , 4) #create a new corrected price column
      corrected_price_cell.value=corrected_price


   #create a object of the class reference
   #this will refer to the 4th column in the excel file
   values=Reference(sheet,
                  min_row=2
                  ,max_row=sheet.max_row
                  ,min_col=4,max_col=4)

   #create a chart and give the data 
   chart=BarChart()
   chart.add_data(values)
   sheet.add_chart(chart,'e2') #the second argument is where we want to add the chart

   #saves the updates, in this case we created a new file to not overwrite the initial file
   wb.save(filename)

process_workbook('transactions.xlsx')


print("aaa")